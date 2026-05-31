from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT.parent / "Credit_Risk_Model_LoanDefaults-main"
OUTPUT_DIR = ROOT / "outputs"

PD_TABLE = {
    "AAA": 0.001,
    "AA": 0.002,
    "A": 0.007,
    "BBB": 0.022,
    "BB": 0.063,
    "B": 0.152,
}

LGD_TABLE = {
    "AAA": 0.20,
    "AA": 0.25,
    "A": 0.35,
    "BBB": 0.45,
    "BB": 0.55,
    "B": 0.65,
}

SCENARIOS = {
    "Baseline": {"AAA": 0.001, "AA": 0.002, "A": 0.007, "BBB": 0.022, "BB": 0.063, "B": 0.152},
    "Downside": {"AAA": 0.003, "AA": 0.008, "A": 0.020, "BBB": 0.055, "BB": 0.130, "B": 0.280},
    "Severe": {"AAA": 0.010, "AA": 0.025, "A": 0.060, "BBB": 0.140, "BB": 0.290, "B": 0.480},
}

RATING_ORDER = ["AAA", "AA", "A", "BBB", "BB", "B"]
SECTORS = ["Retail", "MSME", "Corporate", "Real Estate", "Agri"]


@dataclass(frozen=True)
class BuildConfig:
    seed: int = 42
    loan_count: int = 10_000
    capital_buffer: float = 1.25


def inr_cr(value: float) -> float:
    return value / 10_000_000


def load_lendingclub_recovery_benchmark(source_dir: Path = SOURCE_DIR) -> pd.DataFrame:
    """Summarise real default-only LendingClub recovery data from the provided repo."""
    path = source_dir / "loan_data_defaults.csv"
    columns = ["grade", "loan_amnt", "int_rate", "recovery_rate", "CCF"]
    if not path.exists():
        return pd.DataFrame(columns=["grade", "defaulted_loans", "avg_loan", "avg_int_rate", "empirical_lgd", "avg_ccf"])

    df = pd.read_csv(path, usecols=columns)
    benchmark = (
        df.groupby("grade", as_index=False)
        .agg(
            defaulted_loans=("grade", "size"),
            avg_loan=("loan_amnt", "mean"),
            avg_int_rate=("int_rate", "mean"),
            empirical_recovery_rate=("recovery_rate", "mean"),
            avg_ccf=("CCF", "mean"),
        )
        .assign(empirical_lgd=lambda x: 1 - x["empirical_recovery_rate"])
        [["grade", "defaulted_loans", "avg_loan", "avg_int_rate", "empirical_lgd", "avg_ccf"]]
    )
    return benchmark.round(4)


def generate_portfolio(config: BuildConfig = BuildConfig()) -> pd.DataFrame:
    rng = np.random.default_rng(config.seed)
    n = config.loan_count

    ratings = rng.choice(RATING_ORDER, n, p=[0.10, 0.15, 0.20, 0.25, 0.20, 0.10])
    rating_index = pd.Series(ratings).map({rating: idx for idx, rating in enumerate(RATING_ORDER)}).to_numpy()
    sector = rng.choice(SECTORS, n, p=[0.34, 0.24, 0.18, 0.14, 0.10])
    sector_multiplier = pd.Series(sector).map(
        {"Retail": 1.00, "MSME": 1.08, "Corporate": 0.92, "Real Estate": 1.18, "Agri": 1.12}
    ).to_numpy()

    loan_amount = rng.choice([50_000, 100_000, 250_000, 500_000, 1_000_000], n, p=[0.08, 0.14, 0.22, 0.30, 0.26])
    tenure_years = rng.choice([1, 2, 3, 5, 7, 10], n, p=[0.08, 0.13, 0.24, 0.31, 0.15, 0.09])

    rating_rate_base = pd.Series(ratings).map({"AAA": 8.1, "AA": 9.0, "A": 10.6, "BBB": 12.8, "BB": 16.5, "B": 20.0}).to_numpy()
    interest_rate = np.clip(rating_rate_base + rng.normal(0, 0.85, n) + (sector_multiplier - 1) * 2.0, 7.5, 24.0)

    borrower_income = np.maximum(
        300_000,
        rng.lognormal(mean=np.log(1_250_000), sigma=0.58, size=n) * (1.2 - rating_index * 0.06),
    ).astype(int)
    ltv_ratio = np.clip(rng.normal(0.50 + rating_index * 0.065, 0.08, n), 0.35, 0.96)
    dti_ratio = np.clip((loan_amount / np.maximum(borrower_income, 1)) / tenure_years + rng.normal(0.11, 0.05, n), 0.02, 0.75)

    df = pd.DataFrame(
        {
            "loan_id": np.arange(1, n + 1),
            "loan_amount": loan_amount,
            "credit_rating": pd.Categorical(ratings, categories=RATING_ORDER, ordered=True),
            "tenure_years": tenure_years,
            "interest_rate": interest_rate.round(2),
            "sector": sector,
            "borrower_income": borrower_income,
            "ltv_ratio": ltv_ratio.round(3),
            "dti_ratio": dti_ratio.round(3),
        }
    )

    df["PD"] = df["credit_rating"].map(PD_TABLE).astype(float)
    df["LGD"] = df["credit_rating"].map(LGD_TABLE).astype(float)
    df["EAD"] = df["loan_amount"]
    df["Expected_Loss"] = df["PD"] * df["LGD"] * df["EAD"]
    df["risk_weighted_ead"] = df["EAD"] * (1 + df["PD"] * 4 + df["LGD"])
    df["downgrade_risk"] = (
        df["credit_rating"].isin(["BB", "B"])
        & (df["ltv_ratio"] > 0.75)
        & ((df["dti_ratio"] > 0.35) | (df["PD"] > 0.10))
    )
    df["watchlist_reason"] = np.where(
        df["downgrade_risk"],
        "Weak rating with high leverage / repayment pressure",
        "Performing",
    )
    return df


def build_scenario_results(portfolio: pd.DataFrame, config: BuildConfig = BuildConfig()) -> tuple[pd.DataFrame, pd.DataFrame]:
    loan_level = portfolio[["loan_id", "credit_rating", "LGD", "EAD"]].copy()
    rows = []
    for scenario_name, pd_map in SCENARIOS.items():
        pd_col = f"{scenario_name}_PD"
        el_col = f"{scenario_name}_EL"
        loan_level[pd_col] = loan_level["credit_rating"].map(pd_map).astype(float)
        loan_level[el_col] = loan_level[pd_col] * loan_level["LGD"] * loan_level["EAD"]
        total_el = float(loan_level[el_col].sum())
        total_portfolio = float(loan_level["EAD"].sum())
        rows.append(
            {
                "Scenario": scenario_name,
                "Total_Portfolio": total_portfolio,
                "Expected_Loss": total_el,
                "EL_Pct": total_el / total_portfolio,
                "Capital_Required": total_el * config.capital_buffer,
            }
        )
    return pd.DataFrame(rows), loan_level


def build_segment_tables(portfolio: pd.DataFrame, scenario_loan_level: pd.DataFrame) -> dict[str, pd.DataFrame]:
    severe = scenario_loan_level[["loan_id", "Severe_EL"]]
    enriched = portfolio.merge(severe, on="loan_id", how="left")

    rating_summary = (
        enriched.groupby("credit_rating", observed=False)
        .agg(
            loans=("loan_id", "count"),
            ead=("EAD", "sum"),
            avg_pd=("PD", "mean"),
            avg_lgd=("LGD", "mean"),
            baseline_el=("Expected_Loss", "sum"),
            severe_el=("Severe_EL", "sum"),
            watchlist_loans=("downgrade_risk", "sum"),
        )
        .reset_index()
    )
    rating_summary["el_uplift_x"] = rating_summary["severe_el"] / rating_summary["baseline_el"]

    sector_rating = (
        enriched.pivot_table(
            index="sector",
            columns="credit_rating",
            values="Expected_Loss",
            aggfunc="sum",
            observed=False,
            fill_value=0,
        )
        .reindex(columns=RATING_ORDER)
        .reset_index()
    )

    sector_summary = (
        enriched.groupby("sector", as_index=False)
        .agg(
            loans=("loan_id", "count"),
            ead=("EAD", "sum"),
            baseline_el=("Expected_Loss", "sum"),
            severe_el=("Severe_EL", "sum"),
            watchlist_loans=("downgrade_risk", "sum"),
        )
        .sort_values("severe_el", ascending=False)
    )

    watchlist = (
        enriched.loc[enriched["downgrade_risk"]]
        .sort_values(["Severe_EL", "ltv_ratio", "dti_ratio"], ascending=False)
        .head(100)
    )
    return {
        "rating_summary": rating_summary,
        "sector_rating_matrix": sector_rating,
        "sector_summary": sector_summary,
        "watchlist_top100": watchlist,
    }


def write_excel(
    portfolio: pd.DataFrame,
    scenario_df: pd.DataFrame,
    segment_tables: dict[str, pd.DataFrame],
    benchmark: pd.DataFrame,
    output_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    baseline_el = float(scenario_df.loc[scenario_df["Scenario"] == "Baseline", "Expected_Loss"].iloc[0])
    severe_el = float(scenario_df.loc[scenario_df["Scenario"] == "Severe", "Expected_Loss"].iloc[0])
    shortfall = severe_el - baseline_el
    portfolio_value = float(portfolio["EAD"].sum())

    summary = pd.DataFrame(
        [
            ["Portfolio value", portfolio_value, inr_cr(portfolio_value)],
            ["Baseline expected loss", baseline_el, inr_cr(baseline_el)],
            ["Severe expected loss", severe_el, inr_cr(severe_el)],
            ["Capital shortfall", shortfall, inr_cr(shortfall)],
            ["Severe / baseline EL uplift", severe_el / baseline_el, ""],
            ["Downgrade watchlist loans", int(portfolio["downgrade_risk"].sum()), ""],
        ],
        columns=["Metric", "Value_INR", "Value_Cr"],
    )

    assumptions = pd.DataFrame(
        {
            "credit_rating": RATING_ORDER,
            "PD": [PD_TABLE[x] for x in RATING_ORDER],
            "LGD": [LGD_TABLE[x] for x in RATING_ORDER],
            "Baseline_PD": [SCENARIOS["Baseline"][x] for x in RATING_ORDER],
            "Downside_PD": [SCENARIOS["Downside"][x] for x in RATING_ORDER],
            "Severe_PD": [SCENARIOS["Severe"][x] for x in RATING_ORDER],
        }
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Management_Summary", index=False)
        assumptions.to_excel(writer, sheet_name="Assumptions", index=False)
        scenario_df.to_excel(writer, sheet_name="Scenario_Stress", index=False)
        segment_tables["rating_summary"].to_excel(writer, sheet_name="Rating_Summary", index=False)
        segment_tables["sector_summary"].to_excel(writer, sheet_name="Sector_Summary", index=False)
        segment_tables["sector_rating_matrix"].to_excel(writer, sheet_name="Sector_Rating_Matrix", index=False)
        segment_tables["watchlist_top100"].to_excel(writer, sheet_name="Watchlist_Top100", index=False)
        portfolio.to_excel(writer, sheet_name="Portfolio", index=False)
        benchmark.to_excel(writer, sheet_name="LC_Recovery_Benchmark", index=False)

        workbook = writer.book
        dashboard = workbook.create_sheet("Excel_Dashboard", 0)
        dashboard["A1"] = "Credit Risk Portfolio Stress Test"
        dashboard["A1"].font = Font(size=18, bold=True, color="13233A")
        dashboard["A3"] = "Portfolio Rs Cr"
        dashboard["B3"] = inr_cr(portfolio_value)
        dashboard["A4"] = "Baseline EL Rs Cr"
        dashboard["B4"] = inr_cr(baseline_el)
        dashboard["A5"] = "Severe EL Rs Cr"
        dashboard["B5"] = inr_cr(severe_el)
        dashboard["A6"] = "Capital Shortfall Rs Cr"
        dashboard["B6"] = inr_cr(shortfall)
        dashboard["A7"] = "Severe / Baseline Uplift"
        dashboard["B7"] = severe_el / baseline_el
        for row in range(3, 8):
            dashboard[f"A{row}"].font = Font(bold=True, color="FFFFFF")
            dashboard[f"A{row}"].fill = PatternFill("solid", fgColor="176B87")
            dashboard[f"B{row}"].number_format = "#,##0.00"

        scenario_chart = BarChart()
        scenario_chart.title = "Expected Loss by Scenario"
        scenario_chart.y_axis.title = "INR"
        scenario_chart.x_axis.title = "Scenario"
        scenario_data = Reference(workbook["Scenario_Stress"], min_col=3, min_row=1, max_row=4)
        scenario_cats = Reference(workbook["Scenario_Stress"], min_col=1, min_row=2, max_row=4)
        scenario_chart.add_data(scenario_data, titles_from_data=True)
        scenario_chart.set_categories(scenario_cats)
        scenario_chart.height = 8
        scenario_chart.width = 13
        dashboard.add_chart(scenario_chart, "D3")

        rating_chart = BarChart()
        rating_chart.type = "bar"
        rating_chart.title = "Baseline vs Severe EL by Rating"
        rating_chart.y_axis.title = "Rating"
        rating_chart.x_axis.title = "INR"
        rating_data = Reference(workbook["Rating_Summary"], min_col=6, max_col=7, min_row=1, max_row=7)
        rating_cats = Reference(workbook["Rating_Summary"], min_col=1, min_row=2, max_row=7)
        rating_chart.add_data(rating_data, titles_from_data=True)
        rating_chart.set_categories(rating_cats)
        rating_chart.height = 8
        rating_chart.width = 13
        dashboard.add_chart(rating_chart, "D20")

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.style = "Headline 3"
            for col in sheet.columns:
                max_len = min(max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2, 36)
                sheet.column_dimensions[col[0].column_letter].width = max(12, max_len)

        for sheet_name in ["Management_Summary", "Scenario_Stress", "Rating_Summary", "Sector_Summary"]:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "#,##0.00"
                    if isinstance(cell.value, int):
                        cell.number_format = "#,##0"


def write_dashboard_preview(scenario_df: pd.DataFrame, rating_summary: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    baseline_el = float(scenario_df.loc[scenario_df["Scenario"] == "Baseline", "Expected_Loss"].iloc[0])
    severe_el = float(scenario_df.loc[scenario_df["Scenario"] == "Severe", "Expected_Loss"].iloc[0])
    total = float(scenario_df["Total_Portfolio"].iloc[0])
    shortfall = severe_el - baseline_el

    image = Image.new("RGB", (1600, 950), "white")
    draw = ImageDraw.Draw(image)
    title_font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 36)
    header_font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial Bold.ttf", 24)
    label_font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 20)
    small_font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 17)

    draw.text((44, 36), "Credit Risk Portfolio Stress Test", fill="#13233a", font=title_font)
    draw.text(
        (44, 88),
        f"Portfolio Rs {inr_cr(total):,.1f} Cr | Severe capital shortfall Rs {inr_cr(shortfall):,.1f} Cr",
        fill="#44546a",
        font=label_font,
    )
    draw.text(
        (44, 126),
        "Baseline EL is concentrated in BB/B borrowers; severe stress increases provisions and capital demand.",
        fill="#44546a",
        font=small_font,
    )

    kpis = [
        ("Portfolio", inr_cr(total), "#176B87"),
        ("Baseline EL", inr_cr(baseline_el), "#2A9D8F"),
        ("Severe EL", inr_cr(severe_el), "#E76F51"),
        ("Shortfall", inr_cr(shortfall), "#B00020"),
    ]
    for i, (label, value, color) in enumerate(kpis):
        x0 = 44 + i * 382
        draw.rounded_rectangle((x0, 184, x0 + 335, 300), radius=10, fill=color)
        draw.text((x0 + 22, 204), label, fill="white", font=label_font)
        draw.text((x0 + 22, 244), f"Rs {value:,.1f} Cr", fill="white", font=header_font)

    def draw_bar_chart(x0: int, y0: int, width: int, height: int, labels: list[str], values: list[float], colors: list[str], title: str) -> None:
        draw.rectangle((x0, y0, x0 + width, y0 + height), outline="#d9e2ec", width=2)
        draw.text((x0 + 24, y0 + 18), title, fill="#13233a", font=header_font)
        chart_x0, chart_y0 = x0 + 70, y0 + 78
        chart_w, chart_h = width - 115, height - 145
        draw.line((chart_x0, chart_y0 + chart_h, chart_x0 + chart_w, chart_y0 + chart_h), fill="#6b7280", width=2)
        draw.line((chart_x0, chart_y0, chart_x0, chart_y0 + chart_h), fill="#6b7280", width=2)
        max_value = max(values) * 1.12
        bar_gap = chart_w / len(values)
        bar_w = min(72, bar_gap * 0.55)
        for idx, (label, value, color) in enumerate(zip(labels, values, colors, strict=True)):
            bx = chart_x0 + idx * bar_gap + (bar_gap - bar_w) / 2
            bh = chart_h * value / max_value
            by = chart_y0 + chart_h - bh
            draw.rectangle((bx, by, bx + bar_w, chart_y0 + chart_h), fill=color)
            draw.text((bx - 10, by - 28), f"{value:,.1f}", fill="#13233a", font=small_font)
            draw.text((bx - 8, chart_y0 + chart_h + 16), label, fill="#44546a", font=small_font)

    draw_bar_chart(
        44,
        350,
        710,
        520,
        scenario_df["Scenario"].tolist(),
        scenario_df["Expected_Loss"].map(inr_cr).round(2).tolist(),
        ["#2A9D8F", "#F4A261", "#E76F51"],
        "Expected Loss by Scenario (Rs Cr)",
    )

    draw.rectangle((820, 350, 1546, 870), outline="#d9e2ec", width=2)
    draw.text((844, 368), "Expected Loss by Credit Rating (Rs Cr)", fill="#13233a", font=header_font)
    chart_x0, chart_y0, chart_w, chart_h = 890, 438, 600, 350
    draw.line((chart_x0, chart_y0 + chart_h, chart_x0 + chart_w, chart_y0 + chart_h), fill="#6b7280", width=2)
    draw.line((chart_x0, chart_y0, chart_x0, chart_y0 + chart_h), fill="#6b7280", width=2)
    base_values = rating_summary["baseline_el"].map(inr_cr).round(2).tolist()
    severe_values = rating_summary["severe_el"].map(inr_cr).round(2).tolist()
    max_value = max(severe_values) * 1.15
    group_gap = chart_w / len(RATING_ORDER)
    bar_w = 30
    for idx, rating in enumerate(rating_summary["credit_rating"].astype(str).tolist()):
        gx = chart_x0 + idx * group_gap + group_gap / 2
        for offset, value, color in [(-bar_w, base_values[idx], "#457B9D"), (4, severe_values[idx], "#E63946")]:
            bh = chart_h * value / max_value
            draw.rectangle((gx + offset, chart_y0 + chart_h - bh, gx + offset + bar_w, chart_y0 + chart_h), fill=color)
        draw.text((gx - 26, chart_y0 + chart_h + 16), rating, fill="#44546a", font=small_font)
    draw.rectangle((1040, 820, 1064, 838), fill="#457B9D")
    draw.text((1072, 817), "Baseline", fill="#44546a", font=small_font)
    draw.rectangle((1180, 820, 1204, 838), fill="#E63946")
    draw.text((1212, 817), "Severe", fill="#44546a", font=small_font)

    image.save(output_path)


def run_build(config: BuildConfig = BuildConfig()) -> dict[str, Path | float | int]:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    portfolio = generate_portfolio(config)
    scenario_df, scenario_loan_level = build_scenario_results(portfolio, config)
    segment_tables = build_segment_tables(portfolio, scenario_loan_level)
    benchmark = load_lendingclub_recovery_benchmark()

    portfolio_csv = OUTPUT_DIR / "credit_risk_portfolio.csv"
    scenario_csv = OUTPUT_DIR / "scenario_stress_results.csv"
    workbook_path = OUTPUT_DIR / "credit_risk_powerbi_dataset.xlsx"
    preview_path = OUTPUT_DIR / "dashboard_preview.png"

    portfolio.to_csv(portfolio_csv, index=False)
    scenario_df.to_csv(scenario_csv, index=False)
    write_excel(portfolio, scenario_df, segment_tables, benchmark, workbook_path)
    write_dashboard_preview(scenario_df, segment_tables["rating_summary"], preview_path)

    baseline_el = float(scenario_df.loc[scenario_df["Scenario"] == "Baseline", "Expected_Loss"].iloc[0])
    severe_el = float(scenario_df.loc[scenario_df["Scenario"] == "Severe", "Expected_Loss"].iloc[0])
    total = float(portfolio["EAD"].sum())
    return {
        "portfolio_csv": portfolio_csv,
        "scenario_csv": scenario_csv,
        "workbook": workbook_path,
        "preview": preview_path,
        "portfolio_cr": inr_cr(total),
        "baseline_el_cr": inr_cr(baseline_el),
        "severe_el_cr": inr_cr(severe_el),
        "capital_shortfall_cr": inr_cr(severe_el - baseline_el),
        "el_uplift_x": severe_el / baseline_el,
        "watchlist_loans": int(portfolio["downgrade_risk"].sum()),
    }


if __name__ == "__main__":
    results = run_build()
    for key, value in results.items():
        print(f"{key}: {value}")
